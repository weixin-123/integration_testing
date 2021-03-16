# !/use/bin/env python
# -*- coding:utf-8 -*-

"""
@author:lijian
@describe:
@time:2019/5/30 9:38
"""
# openpyxl只支持 .xlsx
from openpyxl import load_workbook


class DoExcel:
    @classmethod
    def url_replace(cls, path):
        """

        :param path: 路径替换
        :return:
        """
        # 将测试用例的路径替换成同层级下的excel表路径
        file_url = path.replace("test_case", "test_case_excel").replace(".py", ".xlsx")
        return file_url

    @classmethod
    def get_data(cls, file_name, sheet_name):
        excel_url = DoExcel().url_replace(file_name)
        print("excel路径为：", excel_url)
        wb = load_workbook(excel_url, False)
        sheet = wb[sheet_name]

        test_data = []
        for i in range(2, sheet.max_row + 1):
            # sub_data = {}为的是将不同的数据存入到不同的字典中
            sub_data = {}
            sub_data['case_id'] = case_id = sheet.cell(i, 1).value
            sub_data['title'] = module = sheet.cell(i, 2).value
            sub_data['url'] = url = sheet.cell(i, 3).value
            sub_data['data'] = data = sheet.cell(i, 4).value
            sub_data['expected'] = expected = sheet.cell(i, 5).value
            sub_data['result'] = method = sheet.cell(i, 6).value
            test_data.append(sub_data)  # 存储了所有的数据
        return test_data


if __name__ == '__main__':
    print(DoExcel().get_data(r"D:\PycharmProjects\api_test\test_case\test_turntable_express_edit.py", 'Sheet1'))


